import os
import zipfile
from pypdf import PdfReader
from paths import ARC_DIR
from openpyxl import load_workbook
import csv

def test_read_file_xlsx():
    with (zipfile.ZipFile(os.path.join(ARC_DIR, 'file_zip.zip')) as zip_file):
        with zip_file.open('ATTS_column.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=4, column=1).value == 653487

def test_read_file_pdf():
    with (zipfile.ZipFile(os.path.join(ARC_DIR, 'file_zip.zip')) as zip_file):
        with zip_file.open('Lesson1.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            print(len(reader.pages))

            assert len(reader.pages) == 6


def test_read_file_csv():
    with (zipfile.ZipFile(os.path.join(ARC_DIR, 'file_zip.zip')) as zip_file):
        with zip_file.open("columns_ sheet1.csv") as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            third_row = csvreader[2]

            assert third_row[1] == '2'
            assert third_row[3] == '79'